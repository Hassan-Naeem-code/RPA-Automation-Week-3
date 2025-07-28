"""
Data Update Module for RPA Inventory Management System

This module handles saving processed inventory data to various formats and
potentially updating external systems through APIs.

Author: Hassan Naeem
Date: July 2025
"""

import pandas as pd
import json
import logging
from pathlib import Path
from typing import Dict, List, Any, Optional, Union
from datetime import datetime
import requests
import os
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils.dataframe import dataframe_to_rows

# Configure logging
logger = logging.getLogger(__name__)


class InventoryUpdater:
    """
    Handles updating and saving processed inventory data to various destinations.

    Supports CSV, Excel, JSON formats and can integrate with external APIs.
    """

    def __init__(self, config: Optional[Dict[str, Any]] = None):
        """
        Initialize the inventory updater.

        Args:
            config: Configuration dictionary with update parameters
        """
        self.config = config or {}
        self.api_url = self.config.get("api_url")
        self.api_key = self.config.get("api_key")
        self.timeout = self.config.get("timeout_seconds", 30)
        self.max_retries = self.config.get("max_retries", 3)

        logger.info("InventoryUpdater initialized")

    def save_to_csv(self, df: pd.DataFrame, file_path: str, include_index: bool = False) -> bool:
        """
        Save inventory data to CSV file.

        Args:
            df: Processed inventory DataFrame
            file_path: Output file path
            include_index: Whether to include DataFrame index

        Returns:
            True if successful, False otherwise
        """
        try:
            logger.info(f"Saving {len(df)} records to CSV: {file_path}")

            # Ensure output directory exists
            Path(file_path).parent.mkdir(parents=True, exist_ok=True)

            # Save to CSV
            df.to_csv(file_path, index=include_index)

            logger.info(f"Successfully saved data to {file_path}")
            return True

        except Exception as e:
            logger.error(f"Error saving to CSV: {e}")
            return False

    def save_to_excel(
        self, df: pd.DataFrame, file_path: str, sheet_name: str = "Inventory", format_output: bool = True
    ) -> bool:
        """
        Save inventory data to Excel file with optional formatting.

        Args:
            df: Processed inventory DataFrame
            file_path: Output file path
            sheet_name: Name of the Excel sheet
            format_output: Whether to apply formatting

        Returns:
            True if successful, False otherwise
        """
        try:
            logger.info(f"Saving {len(df)} records to Excel: {file_path}")

            # Ensure output directory exists
            Path(file_path).parent.mkdir(parents=True, exist_ok=True)

            if format_output:
                # Create formatted Excel file
                wb = Workbook()
                ws = wb.active
                ws.title = sheet_name

                # Add data to worksheet
                for r in dataframe_to_rows(df, index=False, header=True):
                    ws.append(r)

                # Format headers
                header_font = Font(bold=True, color="FFFFFF")
                header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")

                for cell in ws[1]:
                    cell.font = header_font
                    cell.fill = header_fill
                    cell.alignment = Alignment(horizontal="center")

                # Format stock status column with colors
                if "StockStatus" in df.columns:
                    status_col = None
                    for idx, col in enumerate(df.columns, 1):
                        if col == "StockStatus":
                            status_col = idx
                            break

                    if status_col:
                        for row in range(2, len(df) + 2):
                            cell = ws.cell(row=row, column=status_col)
                            status = cell.value

                            if status == "Critical" or status == "Out of Stock":
                                cell.fill = PatternFill(start_color="FF6B6B", end_color="FF6B6B", fill_type="solid")
                            elif status == "Low Stock":
                                cell.fill = PatternFill(start_color="FFE66D", end_color="FFE66D", fill_type="solid")
                            elif status == "Normal":
                                cell.fill = PatternFill(start_color="4DABF7", end_color="4DABF7", fill_type="solid")

                # Auto-adjust column widths
                for column in ws.columns:
                    max_length = 0
                    column_letter = column[0].column_letter
                    for cell in column:
                        try:
                            if len(str(cell.value)) > max_length:
                                max_length = len(str(cell.value))
                        except:
                            pass
                    adjusted_width = min(max_length + 2, 50)
                    ws.column_dimensions[column_letter].width = adjusted_width

                wb.save(file_path)
            else:
                # Simple Excel save without formatting
                df.to_excel(file_path, sheet_name=sheet_name, index=False)

            logger.info(f"Successfully saved formatted Excel file to {file_path}")
            return True

        except Exception as e:
            logger.error(f"Error saving to Excel: {e}")
            return False

    def save_to_json(self, df: pd.DataFrame, file_path: str, orient: str = "records", indent: int = 2) -> bool:
        """
        Save inventory data to JSON file.

        Args:
            df: Processed inventory DataFrame
            file_path: Output file path
            orient: JSON orientation ('records', 'index', 'values', etc.)
            indent: JSON indentation level

        Returns:
            True if successful, False otherwise
        """
        try:
            logger.info(f"Saving {len(df)} records to JSON: {file_path}")

            # Ensure output directory exists
            Path(file_path).parent.mkdir(parents=True, exist_ok=True)

            # Convert DataFrame to JSON
            json_data = df.to_json(orient=orient, date_format="iso", indent=indent)

            # Write to file
            with open(file_path, "w") as f:
                f.write(json_data)

            logger.info(f"Successfully saved data to {file_path}")
            return True

        except Exception as e:
            logger.error(f"Error saving to JSON: {e}")
            return False

    def save_summary_report(
        self, summary_stats: Dict[str, Any], violations: List[Dict[str, Any]], file_path: str
    ) -> bool:
        """
        Save processing summary and violations to JSON report.

        Args:
            summary_stats: Summary statistics dictionary
            violations: List of business rule violations
            file_path: Output file path

        Returns:
            True if successful, False otherwise
        """
        try:
            logger.info(f"Saving summary report to: {file_path}")

            # Ensure output directory exists
            Path(file_path).parent.mkdir(parents=True, exist_ok=True)

            report_data = {
                "report_generated_at": datetime.now().isoformat(),
                "summary_statistics": summary_stats,
                "business_rule_violations": violations,
                "violation_count": len(violations),
            }

            with open(file_path, "w") as f:
                json.dump(report_data, f, indent=2, default=str)

            logger.info(f"Successfully saved summary report to {file_path}")
            return True

        except Exception as e:
            logger.error(f"Error saving summary report: {e}")
            return False

    def post_to_api(self, data: Union[Dict[str, Any], List[Dict[str, Any]]], endpoint: str = None) -> bool:
        """
        Post inventory data to external API.

        Args:
            data: Data to post (dictionary or list of dictionaries)
            endpoint: API endpoint (if different from config)

        Returns:
            True if successful, False otherwise
        """
        if not self.api_url and not endpoint:
            logger.warning("No API URL configured, skipping API update")
            return False

        url = endpoint or self.api_url

        headers = {"Content-Type": "application/json", "Accept": "application/json"}

        if self.api_key:
            headers["Authorization"] = f"Bearer {self.api_key}"

        try:
            logger.info(f"Posting data to API: {url}")

            for attempt in range(self.max_retries):
                try:
                    response = requests.post(url, json=data, headers=headers, timeout=self.timeout)

                    if response.status_code in [200, 201, 202]:
                        logger.info(f"Successfully posted data to API (status: {response.status_code})")
                        return True
                    else:
                        logger.warning(f"API returned status {response.status_code}: {response.text}")

                except requests.exceptions.Timeout:
                    logger.warning(f"API request timeout (attempt {attempt + 1}/{self.max_retries})")
                except requests.exceptions.ConnectionError:
                    logger.warning(f"API connection error (attempt {attempt + 1}/{self.max_retries})")

            logger.error(f"Failed to post to API after {self.max_retries} attempts")
            return False

        except Exception as e:
            logger.error(f"Error posting to API: {e}")
            return False

    def backup_data(self, df: pd.DataFrame, backup_dir: str = "backups") -> bool:
        """
        Create a timestamped backup of the inventory data.

        Args:
            df: DataFrame to backup
            backup_dir: Directory for backups

        Returns:
            True if successful, False otherwise
        """
        try:
            # Create backup directory
            backup_path = Path(backup_dir)
            backup_path.mkdir(parents=True, exist_ok=True)

            # Generate timestamp filename
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            backup_file = backup_path / f"inventory_backup_{timestamp}.csv"

            # Save backup
            success = self.save_to_csv(df, str(backup_file))

            if success:
                logger.info(f"Backup created: {backup_file}")

            return success

        except Exception as e:
            logger.error(f"Error creating backup: {e}")
            return False

    def update_inventory(
        self,
        df: pd.DataFrame,
        summary_stats: Dict[str, Any],
        violations: List[Dict[str, Any]],
        output_formats: List[str] = None,
        output_dir: str = "data",
    ) -> Dict[str, bool]:
        """
        Main update method that saves data in multiple formats.

        Args:
            df: Processed inventory DataFrame
            summary_stats: Summary statistics
            violations: Business rule violations
            output_formats: List of formats to save ('csv', 'excel', 'json')
            output_dir: Output directory

        Returns:
            Dictionary showing success status for each operation
        """
        if output_formats is None:
            output_formats = ["csv", "excel", "json"]

        results = {}
        output_path = Path(output_dir)

        logger.info(f"Starting inventory update with formats: {output_formats}")

        # Save in requested formats
        if "csv" in output_formats:
            csv_file = output_path / "inventory_processed.csv"
            results["csv"] = self.save_to_csv(df, str(csv_file))

        if "excel" in output_formats:
            excel_file = output_path / "inventory_processed.xlsx"
            results["excel"] = self.save_to_excel(df, str(excel_file))

        if "json" in output_formats:
            json_file = output_path / "inventory_processed.json"
            results["json"] = self.save_to_json(df, str(json_file))

        # Save summary report
        report_file = output_path / "processing_report.json"
        results["report"] = self.save_summary_report(summary_stats, violations, str(report_file))

        # Create backup
        results["backup"] = self.backup_data(df)

        # Post to API if configured
        if self.api_url:
            api_data = df.to_dict("records")
            results["api"] = self.post_to_api(api_data)

        logger.info(f"Update completed. Results: {results}")
        return results


def update_inventory_data(
    df: pd.DataFrame,
    summary_stats: Dict[str, Any],
    violations: List[Dict[str, Any]],
    config: Optional[Dict[str, Any]] = None,
    output_formats: List[str] = None,
) -> Dict[str, bool]:
    """
    Convenience function for updating inventory data.

    Args:
        df: Processed inventory DataFrame
        summary_stats: Summary statistics
        violations: Business rule violations
        config: Optional configuration dictionary
        output_formats: List of output formats

    Returns:
        Dictionary showing success status for each operation
    """
    updater = InventoryUpdater(config)
    return updater.update_inventory(df, summary_stats, violations, output_formats)


if __name__ == "__main__":
    # Example usage
    logging.basicConfig(level=logging.INFO)

    # Create sample data for testing
    sample_data = {
        "SKU": ["SKU001", "SKU002", "SKU003"],
        "Description": ["Item A", "Item B", "Item C"],
        "Location": ["WH1", "WH2", "WH3"],
        "OnHandQty": [50, 15, 0],
        "ReorderPoint": [25, 30, 20],
        "UnitCost": [10.99, 15.50, 8.99],
        "ReorderQty": [0, 15, 20],
        "StockStatus": ["Normal", "Low Stock", "Out of Stock"],
        "TotalValue": [549.50, 232.50, 0.00],
    }

    df = pd.DataFrame(sample_data)
    summary = {"total_records": 3, "processing_timestamp": datetime.now().isoformat()}
    violations = []

    results = update_inventory_data(df, summary, violations)
    print(f"Update results: {results}")
